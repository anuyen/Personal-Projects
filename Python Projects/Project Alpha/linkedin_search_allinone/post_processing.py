import os
import json
import time
import sys
import re

from datetime import datetime
from openpyxl import Workbook, load_workbook


def read_settings():
    settings = []
    if os.path.exists("settings.json"):
        with open("settings.json", "r", encoding="utf-8") as r:
            text = r.read()
            text = text[1:]
            settings = json.loads(text) 
    return settings


def prepare_processing_arguments(setting):
    raw_folder = setting["raw_folder"]
    filters = setting["new_filters"]
    
    if os.path.exists(raw_folder) == False:
        print("No raw_folder, do nothing...")
        return
    if filters is None or len(filters) == 0:
        print("Not found filter content, do nothing...")
        return

    # only process "include" boolean search filter.
    target_rules = [rule for rule in filters["rules"] if " = " in rule]
    process_dict = {}
    process_rules = {}
    for rule in target_rules:
        column = rule.split(' = ')[0]
        keywords = rule.split(' = ')[1]
        keywords = keywords.replace("or", "").replace("and", "")
        keywords = keywords.replace("(", "").replace(")", "")
        keywords = [t.strip().strip("'").lower() for t in keywords.split(' ') if len(t.strip()) > 0]

        if len(column) > 0 and len(keywords) > 0:
            process_rules[column] = list(set(keywords))
            
    if len(process_rules) > 0:
        process_dict["run"] = setting["run"]
        process_dict["folder"] = raw_folder
        process_dict["rules"] = process_rules

    return process_dict


def process_data():
    start_date = datetime.now()
    print("Post proccessing started at {0}.".format(start_date))
    settings = read_settings()
    for setting in settings:
        process_dict = prepare_processing_arguments(setting)
        if len(process_dict) > 0:
            process_keywords(process_dict)

    end_date = datetime.now()
    cost = (end_date - start_date)
    print("Post processing ended at {0}, time cost {1}.".format(end_date, cost))

def process_keywords(process_dict):
    folder = process_dict["folder"]
    column_dict = process_dict["rules"]
    run = process_dict["run"]
    
    spliter = "\\" if "win32" in sys.platform else "/"
    target_files = ["{0}{1}{2}".format(folder, spliter, t) for t in os.listdir(folder) if t.endswith(".json")]
    for file in target_files:
        print("Start process file {0}.".format(file))
        data_dict = {}
        with open(file, "r", encoding="utf8") as r:
            data_dict = json.loads(r.read())
        new_data = []
        if type(data_dict) is list and len(data_dict) > 0:
            for data_group in data_dict:
                new_group = []
                for group_item in data_group:
                    people = group_item[0]
                    if len(people) > 0:
                        for c, keywords in column_dict.items():
                            value = "" if c not in people else people[c].lower()
                            if value is None or len(value) == 0:
                                continue
                            target_keywords = [t for t in keywords if t in value]
                            target_keywords = sorted(target_keywords, key=lambda x: len(x), reverse=True)
                            process_result = []
                            for keyword in target_keywords:
                                    good = False
                                    prefix_reg = "{0}[^a-zéèçàíáóúñ]+".format(keyword)
                                    suffix_reg = "[^a-zéèçàíáóúñ]+{0}".format(keyword)
                                    middle_reg = "[^a-zéèçàíáóúñ]+{0}[^a-zéèçàíáóúñ]+".format(keyword)
                                    if value.startswith(keyword) and re.search(prefix_reg, value) is not None:
                                        good = True
                                    elif value.endswith(keyword) and re.search(suffix_reg, value) is not None:
                                        good = True
                                    elif (value.startswith(keyword) == False and value.endswith(keyword) == False) and re.search(middle_reg, value) is not None:
                                        good = True
                                    elif value == keyword:
                                        good = True

                                    if not good:
                                        process_result.append(False)
                                    else:
                                        process_result.append(True)
                                        break
                            if not (any(process_result)):
                                print("Skip -- keywords: {0}, column: {1}, value: {2}.".format("|".join(target_keywords), c, value))
                                continue
                            else:
                                group_item[0]["work_email"] = process_email(people)
                                new_group.append(group_item)
                new_data.append(new_group)

        filter_file = file.replace(".json", "_filtered.json")
        with open(filter_file, "w", encoding="utf8") as w:
            w.write(json.dumps(new_data))

        grp_idx = 0
        for data_group in new_data:
            grp_idx = grp_idx + 1
            grp_file_name = "{0}_final_{1}_filtered.json".format(run, grp_idx)
            #grp_file_name = filter_file.replace(".json", "_{0}.json".format(grp_idx)).split(spliter)[-1]
            print("Convert the data after post processing to excel format. Saved to {0}.".format(grp_file_name))
            convert_results(grp_file_name, data_group)

def process_email(people_dict):
    new_email = ""

    website = "" if "job_company_website" not in people_dict else people_dict["job_company_website"]
    if website is None or len(website) == 0:
        print(">>> Not found company website, do nothing...")
        return people_dict["work_email"]
    website = website.lower()
    work_email = "" if "work_email" not in people_dict else people_dict["work_email"]
    personal_email = "" if "personal_email" not in people_dict else people_dict["personal_email"]
    current_professional_email = "" if "current_professional_email" not in people_dict else people_dict["current_professional_email"]
    professional_email = "" if "professional_email" not in people_dict else people_dict["professional_email"]
    none_type_email = "" if "none_type_email" not in people_dict else people_dict["none_type_email"]
    work_email_empty = work_email is None or len(work_email) == 0
    personal_email_empty = personal_email is None or len(personal_email) == 0
    current_prof_email_empty = current_professional_email is None or len(current_professional_email) == 0
    professional_email_empty = professional_email is None or len(professional_email) == 0
    none_type_email_empty = none_type_email is None or len(none_type_email) == 0
    empty_list = [work_email_empty, personal_email_empty, current_prof_email_empty, professional_email_empty, none_type_email_empty]
    if all(empty_list):
        print(">>> All emails are empty, do nothing...")
        return people_dict["work_email"]
    
    if not work_email_empty:
        domain = work_email.split('@')[-1].lower()
        if domain in website:
            print("Work email {0} is good match {1}. Don't need change!".format(work_email, website))
            return work_email
        else:
            if not personal_email_empty:
                domain = personal_email.split('@')[-1].lower()
                if domain in website:
                    print("Change work email from {0} to {1}".format(work_email, personal_email))
                    return personal_email
            if not current_prof_email_empty:
                domain = current_professional_email.split('@')[-1].lower()
                if domain in website:
                    print("Change work email from {0} to {1}".format(work_email, current_professional_email))
                    return current_professional_email
            if not professional_email_empty:
                domain = professional_email.split('@')[-1].lower()
                if domain in website:
                    print("Change work email from {0} to {1}".format(work_email, professional_email))
                    return professional_email
            if not none_type_email_empty:
                domain = none_type_email.split('@')[-1].lower()
                if domain in website:
                    print("Change work email from {0} to {1}".format(work_email, none_type_email))
                    return none_type_email
    else:
        if not personal_email_empty:
            domain = personal_email.split('@')[-1].lower()
            if domain in website:
                print("Change work email from {0} to {1}".format(work_email, personal_email))
                return personal_email
        if not current_prof_email_empty:
            domain = current_professional_email.split('@')[-1].lower()
            if domain in website:
                print("Change work email from {0} to {1}".format(work_email, current_professional_email))
                return current_professional_email
        if not professional_email_empty:
            domain = professional_email.split('@')[-1].lower()
            if domain in website:
                print("Change work email from {0} to {1}".format(work_email, professional_email))
                return professional_email
        if not none_type_email_empty:
            domain = none_type_email.split('@')[-1].lower()
            if domain in website:
                print("Change work email from {0} to {1}".format(work_email, none_type_email))
                return none_type_email
    return people_dict["work_email"]


def convert_results(file_name, result_list):

    if len(result_list) == 0:
        return

    output_file_name = file_name.replace(".json", ".xlsx")
    if os.path.exists(output_file_name):
        os.remove(output_file_name)

    wb = Workbook()
    wb.create_sheet("people", index=0)
    people_sheet = wb.get_sheet_by_name("people")
    wb.create_sheet("education", index=1)
    edu_sheet = wb.get_sheet_by_name("education")
    wb.create_sheet("experience", index=2)
    exp_sheet = wb.get_sheet_by_name("experience")

    results = result_list[0]
    people_headers = list(results[0].keys())
    people_sheet.append(people_headers)
    if len(results[1]) > 0:
        exp_headers = list(results[1][0].keys())
        exp_sheet.append(exp_headers)
    if len(results[2]) > 0:
        edu_headers = list(results[2][0].keys())
        edu_sheet.append(edu_headers)

    for results in result_list:
        append_a_line(wb, people_sheet, exp_sheet, edu_sheet, results)
    wb.save(output_file_name)


def append_a_line(wb, people_sheet, exp_sheet, edu_sheet, results):
    
    people_headers = list(results[0].keys())
    # append data to people sheet.
    people_cells = []
    for header in people_headers:
        value = results[0][header]
        if value is None:
            people_cells.append("")
        elif type(value) is int or type(value) is bool or type(value) is str:
            people_cells.append(value)
        else:
            people_cells.append(str(value))
    try:
        people_sheet.append(people_cells)
    except:
        pass

    # append data to exp sheet.
    if len(results[1]) > 0:
        exp_list = results[1]
        exp_headers = list(exp_list[0].keys())
        for exp_dict in exp_list:
            exp_cells = []
            for header in exp_headers:
                value = exp_dict[header]
                if value is None:
                    exp_cells.append("")
                elif type(value) is int or type(value) is bool or type(value) is str:
                    exp_cells.append(value)
                else:
                    exp_cells.append(str(value))
            try:
                exp_sheet.append(exp_cells)
            except:
                pass

    # append data to edu sheet.
    if len(results[2]) > 0:
        edu_list = results[2]
        edu_headers = list(edu_list[0].keys())
        for edu_dict in edu_list:
            edu_cells = []
            for header in edu_headers:
                value = edu_dict[header]
                if value is None:
                    edu_cells.append("")
                elif type(value) is int or type(value) is bool or type(value) is str:
                    edu_cells.append(value)
                else:
                    edu_cells.append(str(value))
            try:
                edu_sheet.append(edu_cells)
            except:
                pass


process_data()

