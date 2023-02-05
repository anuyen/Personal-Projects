import os
import json
import time
import jsonpath
import shutil
import sys
import traceback

from openpyxl import Workbook, load_workbook


def convert_a_file(json_file, setting):
    if not os.path.exists(json_file):
        return None

    raw_folder = setting["raw_folder"]
    spliter = "\\" if "win32" in sys.platform else "/"
    raw_result_file = "{0}{1}{2}".format(raw_folder, spliter, json_file.split(spliter)[-1])
    if os.path.exists(raw_result_file):
        try:
            with open(raw_result_file, "r", encoding="utf-8") as r:
                fake_data = json.loads(r.read(), strict=False)
                print("File is already parsed {0}, skip it for now...".format(json_file))
                return None
        except:
            os.remove(raw_result_file)
    else:
        print("{0} doesn't exist...".format(raw_result_file))
        
    result_list = []
    filters = setting["new_filters"]
    if filters is None or len(filters) == 0:
        print("Not found new_filters section, do nothing now...")
        return
    if "fields" not in filters or len(filters["fields"]) == 0:
        print("Not found filters columns in run {0}, do nothing now...", setting["run"])
        return
    columns = [t.strip() for t in filters["fields"].split('|') if len(t.strip()) > 0]
    rule_set = generate_rule_set(setting)
    exclude_rules = generate_exclude_rule(setting)
    rule_set_empty = rule_set is None or len(rule_set) == 0
    exclude_rule_empty = exclude_rules is None or len(exclude_rules) == 0
    if rule_set_empty and exclude_rule_empty:
        print("Not found filter rules for run {0}, do nothing...".format(setting["run"]))
        return

    with open(json_file, "r", encoding="utf-8") as r:
        for l in r:
            l = l.strip()
            if len(l) > 0:
                try:
                    results = convert_a_line(l, setting, rule_set, exclude_rules, columns)
                    if len(results) == 0:
                        continue
                    #append_a_line(json_file, results)

                    result_list.append(results)
                    if len(result_list) % 100 == 0:
                        print("Cached {0} results for file {1} in run {2}.".format(len(result_list), json_file, setting["run"]))
                except:
                    print(traceback.format_exc())
    

    #raw_folder = setting["raw_folder"]
    #spliter = "\\" if "win32" in sys.platform else "/"
    #raw_result_file = "{0}{1}{2}".format(raw_folder, spliter, file_name.split(spliter)[-1])
    with open(raw_result_file, "w", encoding="utf-8") as w:
        w.write(json.dumps(result_list))
    # ignore internal excel files to speed up.
    #print(">>> Converting {0} to excel...".format(json_file))
    #converted_file = json_file.replace(".json", "_{0}.json".format(setting["run"]))
    #convert_results(converted_file, result_list, setting)


def convert_a_line(json_text, setting, rule_set, excelude_rules, columns):
    if json_text is None or len(json_text) == 0:
        print("Converter - json data is empty, do nothing!")
        return ()
    if type(json_text) is not str:
        print("Converter - json data must be store in a string.")
        return ()

    data_json = None
    try:
        data_json = json.loads(json_text, strict=False)
    except:
        print("\r\n\r\n")
        print("Skip this people....")
        print(json_text)
        print("\r\n\r\n")
        return ()

    people_dict = convert_to_people(data_json)
    if people_dict is None:
        return ()
    good_result = filter_by_necessary_columns(people_dict, setting["necessary_people_columns"])
    if not good_result:
        return ()
    good_result = apply_exclude_rules(people_dict, excelude_rules, columns)
    if not good_result:
        return ()
    good_result = True if rule_set == "" else apply_new_filters(people_dict, rule_set, columns)
    if not good_result:
        return ()
    people_dict_new = apply_people_columns(people_dict, setting["people_columns"])
    
    exp_dicts = () if setting["ignore_exp_table"] else convert_to_experience(data_json)
    edu_dicts = () if setting["ignore_edu_table"] else convert_to_education(data_json)
    return (people_dict_new, exp_dicts, edu_dicts)


# generate exclude filters, apply exclude filters first.
def generate_exclude_rule(setting):
    rule_set = []

    if "new_filters" not in setting:
        return rule_set
    filters = setting["new_filters"]
    if len(filters) == 0:
        return rule_set

    rules = [t.lower() for t in filters["rules"] if " ! " in t]
    if len(rules) == 0:
        return rule_set
    
    for rule in rules:
        column  = rule.split('!')[0]
        content = rule.split('!')[-1]
        if content.count(" or ") > 0 and content.count(" and ") > 0:
            print("Please check the rule [{0}]. ['and' & 'or'] cannot be applied in one rule.".format(rule))
            continue
        else:
            if " or " in content:
                content = content.replace(" or ", " and ")
                content = content.replace("' ", "' not in {0} ".format(column))
                rule_set.append(content)
            elif " and " in content:
                content = content.replace("' ", "' in {0} ".format(column))
                content = "not ({0})".format(content)
                rule_set.append(content)

    return rule_set


# generate the rules to apply on people.
def generate_rule_set(setting):
    rule_set = ""

    if "new_filters" not in setting:
        return rule_set
    filters = setting["new_filters"]
    if len(filters) == 0:
        return rule_set
    relation = filters["relation"]
    rules = [t.lower() for t in filters["rules"] if " ! " not in t]
    new_rules = []
    for rule in rules:
        if " = " in rule:
            column = rule.split('=')[0].strip()
            content = rule.split('=')[1].strip()
            content = content.replace("' ", "' in {0} ".format(column))
            new_rules.append(content)
        else:
            new_rules.append(rule)

    rule_set = relation.format(*new_rules)

    return rule_set


# filter by include filter.
def apply_new_filters(data_dict, rule_set, columns):

    #columns = [t.strip() for t in filters["fields"].split('|') if len(t.strip()) > 0]
    column_vars = locals()
    for column in columns:
        cell_value = "" if column not in data_dict else data_dict[column]
        cell_value = "" if cell_value is None else str(cell_value).lower()
        column_vars[column] = cell_value
    return eval(rule_set)


# filter by exclue filter first.
def apply_exclude_rules(data_dict, rule_set, columns):
    #columns = [t.strip() for t in filters["fields"].split('|') if len(t.strip()) > 0]
    column_vars = locals()
    for column in columns:
        cell_value = "" if column not in data_dict else data_dict[column]
        cell_value = "" if cell_value is None else str(cell_value).lower()
        column_vars[column] = cell_value

    for rule in rule_set:
        if eval(rule) == False:
            return False

    return True

def convert_results(file_name, result_list, setting):

    if len(result_list) == 0:
        return

    output_file_name = file_name.replace(".json", ".xlsx")
    if os.path.exists(output_file_name):
        os.remove(output_file_name)

    wb = Workbook()
    wb.create_sheet("people", index=0)
    people_sheet = wb.get_sheet_by_name("people")
    wb.create_sheet("education", index=1)
    edu_sheet = wb.get_sheet_by_name("education")
    wb.create_sheet("experience", index=2)
    exp_sheet = wb.get_sheet_by_name("experience")

    results = result_list[0]
    people_headers = list(results[0].keys())
    people_sheet.append(people_headers)
    if len(results[1]) > 0:
        exp_headers = list(results[1][0].keys())
        exp_sheet.append(exp_headers)
    if len(results[2]) > 0:
        edu_headers = list(results[2][0].keys())
        edu_sheet.append(edu_headers)

    for results in result_list:
        append_a_line(wb, people_sheet, exp_sheet, edu_sheet, results)
    wb.save(output_file_name)


def append_a_line(wb, people_sheet, exp_sheet, edu_sheet, results):
    
    people_headers = list(results[0].keys())
    # append data to people sheet.
    people_cells = []
    for header in people_headers:
        value = results[0][header]
        if value is None:
            people_cells.append("")
        elif type(value) is int or type(value) is bool or type(value) is str:
            people_cells.append(value)
        else:
            people_cells.append(str(value))
    try:
        people_sheet.append(people_cells)
    except:
        pass

    # append data to exp sheet.
    if len(results[1]) > 0:
        exp_list = results[1]
        exp_headers = list(exp_list[0].keys())
        for exp_dict in exp_list:
            exp_cells = []
            for header in exp_headers:
                value = exp_dict[header]
                if value is None:
                    exp_cells.append("")
                elif type(value) is int or type(value) is bool or type(value) is str:
                    exp_cells.append(value)
                else:
                    exp_cells.append(str(value))
            try:
                exp_sheet.append(exp_cells)
            except:
                pass

    # append data to edu sheet.
    if len(results[2]) > 0:
        edu_list = results[2]
        edu_headers = list(edu_list[0].keys())
        for edu_dict in edu_list:
            edu_cells = []
            for header in edu_headers:
                value = edu_dict[header]
                if value is None:
                    edu_cells.append("")
                elif type(value) is int or type(value) is bool or type(value) is str:
                    edu_cells.append(value)
                else:
                    edu_cells.append(str(value))
            try:
                edu_sheet.append(edu_cells)
            except:
                pass


def convert_to_people(data_json):
    column_dict = {}
    id_node = jsonpath.jsonpath(data_json, "$.id")
    column_dict["id"] = "" if id_node == False else id_node[0]
    if column_dict["id"] is None or len(column_dict["id"]) == 0:
        return None
    full_name_node = jsonpath.jsonpath(data_json, "$.full_name")
    column_dict["full_name"] = "" if full_name_node == False else full_name_node[0]
    column_dict["full_name"] = "" if column_dict["full_name"] is None or len(column_dict["full_name"]) == 0 else column_dict["full_name"].capitalize()
    first_name_node = jsonpath.jsonpath(data_json, "$.first_name")
    column_dict["first_name"] = "" if first_name_node == False else first_name_node[0]
    column_dict["first_name"] = "" if column_dict["first_name"] is None or len(column_dict["first_name"]) == 0 else column_dict["first_name"].capitalize()
    middle_initial_node = jsonpath.jsonpath(data_json, "$.middle_initial")
    column_dict["middle_initial"] = "" if middle_initial_node == False else middle_initial_node[0]
    middle_name_node = jsonpath.jsonpath(data_json, "$.middle_name")
    column_dict["middle_name"] = "" if middle_name_node == False else middle_name_node[0]
    column_dict["middle_name"] = "" if column_dict["middle_name"] is None or len(column_dict["middle_name"]) == 0 else column_dict["middle_name"].capitalize()
    last_name_node = jsonpath.jsonpath(data_json, "$.last_name")
    column_dict["last_name"] = "" if last_name_node == False else last_name_node[0]
    column_dict["last_name"] = "" if column_dict["last_name"] is None or len(column_dict["last_name"]) == 0 else column_dict["last_name"].capitalize()
    gender_node = jsonpath.jsonpath(data_json, "$.gender")
    column_dict["gender"] = "" if gender_node == False else gender_node[0]
    birth_year_node = jsonpath.jsonpath(data_json, "$.birth_year")
    column_dict["birth_year"] = "" if birth_year_node == False else birth_year_node[0]
    birth_date_node = jsonpath.jsonpath(data_json, "$.birth_date")
    column_dict["birth_date"] = "" if birth_date_node == False else birth_date_node[0]
    linkedin_url_node = jsonpath.jsonpath(data_json, "$.linkedin_url")
    column_dict["linkedin_url"] = "" if linkedin_url_node == False else linkedin_url_node[0]
    linkedin_username_node = jsonpath.jsonpath(data_json, "$.linkedin_username")
    column_dict["linkedin_username"] = "" if linkedin_username_node == False else linkedin_username_node[0]
    linkedin_id_node = jsonpath.jsonpath(data_json, "$.linkedin_id")
    column_dict["linkedin_id"] = "" if linkedin_id_node == False else linkedin_id_node[0]
    facebook_url_node = jsonpath.jsonpath(data_json, "$.facebook_url")
    column_dict["facebook_url"] = "" if facebook_url_node == False else facebook_url_node[0]
    facebook_username_node = jsonpath.jsonpath(data_json, "$.facebook_username")
    column_dict["facebook_username"] = "" if facebook_username_node == False else facebook_username_node[0]
    facebook_id_node = jsonpath.jsonpath(data_json, "$.facebook_id")
    column_dict["facebook_id"] = "" if facebook_id_node == False else facebook_id_node[0]
    twitter_url_node = jsonpath.jsonpath(data_json, "$.twitter_url")
    column_dict["twitter_url"] = "" if twitter_url_node == False else twitter_url_node[0]
    twitter_username_node = jsonpath.jsonpath(data_json, "$.twitter_username")
    column_dict["twitter_username"] = "" if twitter_username_node == False else twitter_username_node[0]
    github_url_node = jsonpath.jsonpath(data_json, "$.github_url")
    column_dict["github_url"] = "" if github_url_node == False else github_url_node[0]
    github_username_node = jsonpath.jsonpath(data_json, "$.github_username")
    column_dict["github_username"] = "" if github_username_node == False else github_username_node[0]
    work_email_node = jsonpath.jsonpath(data_json, "$.work_email")
    column_dict["work_email"] = "" if work_email_node == False else work_email_node[0]
    mobile_phone_node = jsonpath.jsonpath(data_json, "$.mobile_phone")
    column_dict["mobile_phone"] = "" if mobile_phone_node == False else mobile_phone_node[0]
    industry_node = jsonpath.jsonpath(data_json, "$.industry")
    column_dict["industry"] = "" if industry_node == False else industry_node[0]
    job_title_node = jsonpath.jsonpath(data_json, "$.job_title")
    column_dict["job_title"] = "" if job_title_node == False else job_title_node[0]
    job_title_role_node = jsonpath.jsonpath(data_json, "$.job_title_role")
    column_dict["job_title_role"] = "" if job_title_role_node == False else job_title_role_node[0]
    job_title_sub_role_node = jsonpath.jsonpath(data_json, "$.job_title_sub_role")
    column_dict["job_title_sub_role"] = "" if job_title_sub_role_node == False else job_title_sub_role_node[0]
    job_title_levels_node = jsonpath.jsonpath(data_json, "$.job_title_levels")
    column_dict["job_title_levels"] = "" if job_title_levels_node == False else job_title_levels_node[0]
    job_company_id_node = jsonpath.jsonpath(data_json, "$.job_company_id")
    column_dict["job_company_id"] = "" if job_company_id_node == False else job_company_id_node[0]
    job_company_name_node = jsonpath.jsonpath(data_json, "$.job_company_name")
    column_dict["job_company_name"] = "" if job_company_name_node == False else job_company_name_node[0]
    job_company_website_node = jsonpath.jsonpath(data_json, "$.job_company_website")
    column_dict["job_company_website"] = "" if job_company_website_node == False else job_company_website_node[0]
    job_company_size_node = jsonpath.jsonpath(data_json, "$.job_company_size")
    column_dict["job_company_size"] = "" if job_company_size_node == False else job_company_size_node[0]
    job_company_founded_node = jsonpath.jsonpath(data_json, "$.job_company_founded")
    column_dict["job_company_founded"] = "" if job_company_founded_node == False else job_company_founded_node[0]
    job_company_industry_node = jsonpath.jsonpath(data_json, "$.job_company_industry")
    column_dict["job_company_industry"] = "" if job_company_industry_node == False else job_company_industry_node[0]
    job_company_linkedin_url_node = jsonpath.jsonpath(data_json, "$.job_company_linkedin_url")
    column_dict["job_company_linkedin_url"] = "" if job_company_linkedin_url_node == False else job_company_linkedin_url_node[0]
    job_company_linkedin_id_node = jsonpath.jsonpath(data_json, "$.job_company_linkedin_id")
    column_dict["job_company_linkedin_id"] = "" if job_company_linkedin_id_node == False else job_company_linkedin_id_node[0]
    job_company_facebook_url_node = jsonpath.jsonpath(data_json, "$.job_company_facebook_url")
    column_dict["job_company_facebook_url"] = "" if job_company_facebook_url_node == False else job_company_facebook_url_node[0]
    job_company_twitter_url_node = jsonpath.jsonpath(data_json, "$.job_company_twitter_url")
    column_dict["job_company_twitter_url"] = "" if job_company_twitter_url_node == False else job_company_twitter_url_node[0]
    job_company_location_name_node = jsonpath.jsonpath(data_json, "$.job_company_location_name")
    column_dict["job_company_location_name"] = "" if job_company_location_name_node == False else job_company_location_name_node[0]
    job_company_location_locality_node = jsonpath.jsonpath(data_json, "$.job_company_location_locality")
    column_dict["job_company_location_locality"] = "" if job_company_location_locality_node == False else job_company_location_locality_node[0]
    job_company_location_metro_node = jsonpath.jsonpath(data_json, "$.job_company_location_metro")
    column_dict["job_company_location_metro"] = "" if job_company_location_metro_node == False else job_company_location_metro_node[0]
    job_company_location_region_node = jsonpath.jsonpath(data_json, "$.job_company_location_region")
    column_dict["job_company_location_region"] = "" if job_company_location_region_node == False else job_company_location_region_node[0]
    job_company_location_geo_node = jsonpath.jsonpath(data_json, "$.job_company_location_geo")
    column_dict["job_company_location_geo"] = "" if job_company_location_geo_node == False else job_company_location_geo_node[0]
    job_company_location_street_address_node = jsonpath.jsonpath(data_json, "$.job_company_location_street_address")
    column_dict["job_company_location_street_address"] = "" if job_company_location_street_address_node == False else job_company_location_street_address_node[0]
    job_company_location_address_line_2_node = jsonpath.jsonpath(data_json, "$.job_company_location_address_line_2")
    column_dict["job_company_location_address_line_2"] = "" if job_company_location_address_line_2_node == False else job_company_location_address_line_2_node[0]
    job_company_location_postal_code_node = jsonpath.jsonpath(data_json, "$.job_company_location_postal_code")
    column_dict["job_company_location_postal_code"] = "" if job_company_location_postal_code_node == False else job_company_location_postal_code_node[0]
    job_company_location_country_node = jsonpath.jsonpath(data_json, "$.job_company_location_country")
    column_dict["job_company_location_country"] = "" if job_company_location_country_node == False else job_company_location_country_node[0]
    job_company_location_continent_node = jsonpath.jsonpath(data_json, "$.job_company_location_continent")
    column_dict["job_company_location_continent"] = "" if job_company_location_continent_node == False else job_company_location_continent_node[0]
    job_last_updated_node = jsonpath.jsonpath(data_json, "$.job_last_updated")
    column_dict["job_last_updated"] = "" if job_last_updated_node == False else job_last_updated_node[0]
    job_start_date_node = jsonpath.jsonpath(data_json, "$.job_start_date")
    column_dict["job_start_date"] = "" if job_start_date_node == False else job_start_date_node[0]
    job_summary_node = jsonpath.jsonpath(data_json, "$.job_summary")
    column_dict["job_summary"] = "" if job_summary_node == False else job_summary_node[0]
    location_name_node = jsonpath.jsonpath(data_json, "$.location_name")
    column_dict["location_name"] = "" if location_name_node == False else location_name_node[0]
    location_locality_node = jsonpath.jsonpath(data_json, "$.location_locality")
    column_dict["location_locality"] = "" if location_locality_node == False else location_locality_node[0]
    location_metro_node = jsonpath.jsonpath(data_json, "$.location_metro")
    column_dict["location_metro"] = "" if location_metro_node == False else location_metro_node[0]
    location_region_node = jsonpath.jsonpath(data_json, "$.location_region")
    column_dict["location_region"] = "" if location_region_node == False else location_region_node[0]
    location_country_node = jsonpath.jsonpath(data_json, "$.location_country")
    column_dict["location_country"] = "" if location_country_node == False else location_country_node[0]
    location_continent_node = jsonpath.jsonpath(data_json, "$.location_continent")
    column_dict["location_continent"] = "" if location_continent_node == False else location_continent_node[0]
    location_street_address_node = jsonpath.jsonpath(data_json, "$.location_street_address")
    column_dict["location_street_address"] = "" if location_street_address_node == False else location_street_address_node[0]
    location_address_line_2_node = jsonpath.jsonpath(data_json, "$.location_address_line_2")
    column_dict["location_address_line_2"] = "" if location_address_line_2_node == False else location_address_line_2_node[0]
    location_postal_code_node = jsonpath.jsonpath(data_json, "$.location_postal_code")
    column_dict["location_postal_code"] = "" if location_postal_code_node == False else location_postal_code_node[0]
    location_geo_node = jsonpath.jsonpath(data_json, "$.location_geo")
    column_dict["location_geo"] = "" if location_geo_node == False else location_geo_node[0]
    location_last_updated_node = jsonpath.jsonpath(data_json, "$.location_last_updated")
    column_dict["location_last_updated"] = "" if location_last_updated_node == False else location_last_updated_node[0]
    linkedin_connections_node = jsonpath.jsonpath(data_json, "$.linkedin_connections")
    column_dict["linkedin_connections"] = "" if linkedin_connections_node == False else linkedin_connections_node[0]
    inferred_salary_node = jsonpath.jsonpath(data_json, "$.inferred_salary")
    column_dict["inferred_salary"] = "" if inferred_salary_node == False else inferred_salary_node[0]
    inferred_years_experience_node = jsonpath.jsonpath(data_json, "$.inferred_years_experience")
    column_dict["inferred_years_experience"] = "" if inferred_years_experience_node == False else inferred_years_experience_node[0]
    summary_node = jsonpath.jsonpath(data_json, "$.summary")
    column_dict["summary"] = "" if summary_node == False else summary_node[0]
    phone_numbers_node = jsonpath.jsonpath(data_json, "$.phone_numbers")
    column_dict["phone_numbers"] = "" if phone_numbers_node == False else phone_numbers_node[0]
    emails_node = jsonpath.jsonpath(data_json, "$.emails")
    column_dict["emails"] = "" if emails_node == False else emails_node[0]
    professional_email_node = jsonpath.jsonpath(data_json, "$.emails[?(@.type=='professional')].address")
    column_dict["professional_email"] = "" if professional_email_node == False else professional_email_node[0]
    current_professional_email_node = jsonpath.jsonpath(data_json, "$.emails[?(@.type=='current_professional')].address")
    column_dict["current_professional_email"] = "" if current_professional_email_node == False else current_professional_email_node[0]
    personal_email_node = jsonpath.jsonpath(data_json, "$.emails[?(@.type=='personal')].address")
    column_dict["personal_email"] = "" if personal_email_node == False else personal_email_node[0]
    none_type_email_node = jsonpath.jsonpath(data_json, "$.emails[?(@.type==None)].address")
    column_dict["none_type_email"] = "" if none_type_email_node == False else none_type_email_node[0]
    interests_node = jsonpath.jsonpath(data_json, "$.interests")
    column_dict["interests"] = "" if interests_node == False else interests_node[0]
    skills_node = jsonpath.jsonpath(data_json, "$.skills")
    column_dict["skills"] = "" if skills_node == False else skills_node[0]
    location_names_node = jsonpath.jsonpath(data_json, "$.location_names")
    column_dict["location_names"] = "" if location_names_node == False else location_names_node[0]
    regions_node = jsonpath.jsonpath(data_json, "$.regions")
    column_dict["regions"] = "" if regions_node == False else regions_node[0]
    countries_node = jsonpath.jsonpath(data_json, "$.countries")
    column_dict["countries"] = "" if countries_node == False else countries_node[0]
    street_addresses_node = jsonpath.jsonpath(data_json, "$.street_addresses")
    column_dict["street_addresses"] = "" if street_addresses_node == False else street_addresses_node[0]
    return column_dict


# detect whether the line is good!
def filter_by_necessary_columns(data_dict, necessary_columns):
    good = True
    optional_columns = [t for t in necessary_columns if type(t) is list]
    necessary_columns = [t for t in necessary_columns if type(t) is str]
    for column in necessary_columns:
        if column in data_dict:
            value = data_dict[column]
            if value is None:
                good = False
                break
            else:
                if type(value) is str:
                    value = value.strip()
                    if len(value) == 0:
                        good = False
                        break
                elif type(value) is dict or type(value) is list:
                    if len(value) == 0:
                        good = False
                        break
    option_results = []
    if good:
        for option_list in optional_columns:
            option_good = False
            for option_column in option_list:
                if option_column in data_dict:
                    value = data_dict[option_column]
                    if value is not None:
                        if type(value) is dict or type(value) is list:
                            if len(value) > 0:
                                option_good = True
                                break
                        elif type(value) is str:
                            value = value.strip()
                            if len(value) > 0:
                                option_good = True
                                break          
            option_results.append(option_good)

    good = good and all(option_results)
    return good


# only generate configure columns to output file.
def apply_people_columns(data_dict, columns):
    result_dict = {}
    for column in columns:
        if column in data_dict:
            result_dict[column] = data_dict[column]
    return result_dict


def convert_to_people_sql(column_dict):
    sql_template = "INSERT INTO people ({0}) VALUES ({1});"
    headers = list(column_dict.keys())
    column_cluster = ",".join(column_dict.keys())
    cells = []
    for header in headers:
        if column_dict[header] is None:
            cells.append("null")
        elif type(column_dict[header]) is int:
            cells.append(str(column_dict[header]))
        elif type(column_dict[header]) is str:
            cells.append("'{0}'".format(column_dict[header].replace("'", "''")))
        else:
            cells.append("null")
    value_cluster = ",".join(cells)
    sql = sql_template.format(column_cluster, value_cluster)
    return sql


def convert_to_experience(data_json):
    column_dict = {}
    id_node = jsonpath.jsonpath(data_json, "$.id")
    column_dict["people_id"] = "" if id_node == False else id_node[0]

    exp_nodes = jsonpath.jsonpath(data_json, "$.experience[*]")
    exp_dicts = []
    if exp_nodes != False:
        for exp_node in exp_nodes:
            exp_dict = column_dict.copy()
            company_name_node = jsonpath.jsonpath(exp_node, "$.company.name")
            exp_dict["company_name"] = "" if company_name_node == False else company_name_node[0]
            company_size_node =  jsonpath.jsonpath(exp_node, "$.company.size")
            exp_dict["company_size"] = "" if company_size_node == False else company_size_node[0]
            company_id_node = jsonpath.jsonpath(exp_node, "$.company.id")
            exp_dict["company_id"] = "" if company_id_node == False else company_id_node[0]
            company_founded_node = jsonpath.jsonpath(exp_node, "$.company.founded")
            exp_dict["company_founded"] = "" if company_founded_node == False else company_founded_node[0]
            company_industry_node = jsonpath.jsonpath(exp_node, "$.company.industry")
            exp_dict["company_industry"] = "" if company_industry_node == False else company_industry_node[0]
            company_location_name_node = jsonpath.jsonpath(exp_node, "$.company.location.name")
            exp_dict["company_location_name"] = "" if company_location_name_node == False else company_location_name_node[0]
            company_location_locality_node = jsonpath.jsonpath(exp_node, "$.company.location.locality")
            exp_dict["company_location_locality"] = "" if company_location_locality_node == False else company_location_locality_node[0]
            company_location_region_node = jsonpath.jsonpath(exp_node, "$.company.location.region")
            exp_dict["company_location_region"] = "" if company_location_region_node == False else company_location_region_node[0]
            company_location_metro_node = jsonpath.jsonpath(exp_node, "$.company.location.metro")
            exp_dict["company_location_metro"] = "" if company_location_metro_node == False else company_location_metro_node[0]
            company_location_country_node = jsonpath.jsonpath(exp_node, "$.company.location.country")
            exp_dict["company_location_country"] = "" if company_location_country_node == False else company_location_country_node[0]
            company_location_continent_node = jsonpath.jsonpath(exp_node, "$.company.location.continent")
            exp_dict["company_location_continent"] = "" if company_location_continent_node == False else company_location_continent_node[0]
            company_location_street_address_node = jsonpath.jsonpath(exp_node, "$.company.location.street_address")
            exp_dict["company_location_street_address"] = "" if company_location_street_address_node == False else company_location_street_address_node[0]
            company_location_address_line_2_node = jsonpath.jsonpath(exp_node, "$.company.location.address_line_2")
            exp_dict["company_location_address_line_2"] = "" if company_location_address_line_2_node == False else company_location_address_line_2_node[0]
            company_location_postal_code_node = jsonpath.jsonpath(exp_node, "$.company.location.postal_code")
            exp_dict["company_location_postal_code"] = "" if company_location_postal_code_node == False else company_location_postal_code_node[0]
            company_location_geo_node = jsonpath.jsonpath(exp_node, "$.company.location.geo")
            exp_dict["company_location_geo"] = "" if company_location_geo_node == False else company_location_geo_node[0]
            company_linkedin_url_node = jsonpath.jsonpath(exp_node, "$.company.linkedin_url")
            exp_dict["company_linkedin_url"] = "" if company_linkedin_url_node == False else company_linkedin_url_node[0]
            company_linkedin_id_node = jsonpath.jsonpath(exp_node, "$.company.linkedin_id")
            exp_dict["company_linkedin_id"] = "" if company_linkedin_id_node == False else company_linkedin_id_node[0]
            company_facebook_url_node = jsonpath.jsonpath(exp_node, "$.company.facebook_url")
            exp_dict["company_facebook_url"] = "" if company_facebook_url_node == False else company_facebook_url_node[0]
            company_twitter_url_node = jsonpath.jsonpath(exp_node, "$.company.twitter_url")
            exp_dict["company_twitter_url"] = "" if company_twitter_url_node == False else company_twitter_url_node[0]
            company_website_node = jsonpath.jsonpath(exp_node, "$.company.website")
            exp_dict["company_website"] = "" if company_website_node == False else company_website_node[0]
            location_names_node = jsonpath.jsonpath(exp_node, "$.location_names")
            exp_dict["location_names"] = "" if location_names_node == False else location_names_node[0]
            start_date_node = jsonpath.jsonpath(exp_node, "$.start_date")
            exp_dict["start_date"] = "" if start_date_node == False else start_date_node[0]
            end_date_node = jsonpath.jsonpath(exp_node, "$.end_date")
            exp_dict["end_date"] = "" if end_date_node == False else end_date_node[0]
            title_name_node = jsonpath.jsonpath(exp_node , "$.title.name")
            exp_dict["title_name"] = "" if title_name_node == False else title_name_node[0]
            title_role_node = jsonpath.jsonpath(exp_node, "$.title.role")
            exp_dict["title_role"] = "" if title_role_node == False else title_role_node[0]
            title_sub_role_node = jsonpath.jsonpath(exp_node, "$.title.sub_role")
            exp_dict["title_sub_role"] = "" if title_sub_role_node == False else title_sub_role_node[0]
            title_levels_node = jsonpath.jsonpath(exp_node, "$.title.levels")
            exp_dict["title_levels"] = "" if title_levels_node == False else title_levels_node[0]
            is_primary_node = jsonpath.jsonpath(exp_node, "$.is_primary")
            exp_dict["is_primary"] = "" if is_primary_node == False else is_primary_node[0]
            summary_node = jsonpath.jsonpath(exp_node, "$.summary")
            exp_dict["summary"] = "" if summary_node == False else summary_node[0]
            
            exp_dicts.append(exp_dict)
        
    return exp_dicts


def convert_to_experience_sql(exp_dict_list):
    sqls =[]

    for exp_dict in exp_dict_list:
        sql_template = "INSERT INTO work_experience ({0}) VALUES({1});";
        headers = list(exp_dict.keys())
        column_cluster = ",".join(exp_dict.keys())
        cells = []
        for header in headers:
            if exp_dict[header] is None:
                cells.append("null")
            elif type(exp_dict[header]) is int:
                cells.append(str(exp_dict[header]))
            elif type(exp_dict[header]) is bool:
                cells.append(str(exp_dict[header]))
            elif type(exp_dict[header]) is str:
                cells.append("'{0}'".format(exp_dict[header].replace("'", "''")))
            else:
                cells.append("null")
        value_cluster = ",".join(cells)
        sql = sql_template.format(column_cluster, value_cluster)
        sqls.append(sql)

    return sqls


def convert_to_education(data_json):
    column_dict = {}
    id_node = jsonpath.jsonpath(data_json, "$.id")
    column_dict["people_id"] = "" if id_node == False else id_node[0]

    edu_dicts = []
    education_nodes = jsonpath.jsonpath(data_json, "$.education[*]")
    if education_nodes != False:
        for edu_node in education_nodes:
            edu_dict = column_dict.copy()

            school_name_node = jsonpath.jsonpath(edu_node, "$.school.name")
            edu_dict["school_name"] = "" if school_name_node == False else school_name_node[0]
            school_type_node = jsonpath.jsonpath(edu_node, "$.school.type")
            edu_dict["school_type"] = "" if school_type_node == False else school_type_node[0]
            school_id_node = jsonpath.jsonpath(edu_node, "$.school.id")
            edu_dict["school_id"] = "" if school_id_node == False else school_id_node[0]
            school_location_name_node = jsonpath.jsonpath(edu_node, "$.school.location.name")
            edu_dict["school_location_name"] = "" if school_location_name_node == False else school_location_name_node[0]
            school_location_locality_node = jsonpath.jsonpath(edu_node, "$.school.location.locality")
            edu_dict["school_location_locality"] = "" if school_location_locality_node == False else school_location_locality_node[0]
            school_location_region_node = jsonpath.jsonpath(edu_node, "$.school.location.region")
            edu_dict["school_location_region"] = "" if school_location_region_node == False else school_location_region_node[0]
            school_location_country_node = jsonpath.jsonpath(edu_node, "$.school.location.country")
            edu_dict["school_location_country"] = "" if school_location_country_node == False else school_location_country_node[0]
            school_location_continent_node = jsonpath.jsonpath(edu_node, "$.school.location.continent")
            edu_dict["school_location_continent"] = "" if school_location_continent_node == False else school_location_continent_node[0]
            school_linkedin_url_node = jsonpath.jsonpath(edu_node, "$.school.linkedin_url")
            edu_dict["school_linkedin_url"] = "" if school_linkedin_url_node == False else school_linkedin_url_node[0]
            school_facebook_url_node = jsonpath.jsonpath(edu_node, "$.school.facebook_url")
            edu_dict["school_facebook_url"] = "" if school_facebook_url_node == False else school_facebook_url_node[0]
            school_twitter_url_node = jsonpath.jsonpath(edu_node, "$.school.twitter_url")
            edu_dict["school_twitter_url"] = "" if school_twitter_url_node == False else school_twitter_url_node[0]
            school_linkedin_id_node = jsonpath.jsonpath(edu_node, "$.school.linkedin_id")
            edu_dict["school_linkedin_id"] = "" if school_linkedin_id_node == False else school_linkedin_id_node[0]
            school_website_node = jsonpath.jsonpath(edu_node, "$.school.website")
            edu_dict["school_website"] = "" if school_website_node == False else school_website_node[0]
            school_domain_node = jsonpath.jsonpath(edu_node, "$.school.domain")
            edu_dict["school_domain"] = "" if school_domain_node == False else school_domain_node[0]
            start_date_node = jsonpath.jsonpath(edu_node, "$.start_date")
            edu_dict["start_date"] = "" if start_date_node == False else start_date_node[0]
            end_date_node = jsonpath.jsonpath(edu_node, "$.end_date")
            edu_dict["end_date"] = "" if end_date_node == False else end_date_node[0]
            degrees_node = jsonpath.jsonpath(edu_node, "$.degrees")
            edu_dict["degrees"] = "" if degrees_node == False else degrees_node[0]
            majors_node = jsonpath.jsonpath(edu_node, "$.majors")
            edu_dict["majors"] = "" if majors_node == False else majors_node[0]
            minors_node = jsonpath.jsonpath(edu_node, "$.minors")
            edu_dict["minors"] = "" if minors_node == False else minors_node[0]
            summary_node = jsonpath.jsonpath(edu_node, "$.summary")
            edu_dict["summary"] = "" if summary_node == False else summary_node[0]

            edu_dicts.append(edu_dict)

    return edu_dicts


def convert_to_education_sql(edu_dicts):
    sqls = []
    for edu_dict in edu_dicts:
        sql_template = "INSERT INTO work_experience ({0}) VALUES({1});";
        headers = list(edu_dict.keys())
        column_cluster = ",".join(edu_dict.keys())
        cells = []
        for header in headers:
            if edu_dict[header] is None:
                cells.append("null")
            elif type(edu_dict[header]) is int:
                cells.append(str(edu_dict[header]))
            elif type(edu_dict[header]) is bool:
                cells.append(str(edu_dict[header]))
            elif type(edu_dict[header]) is str:
                cells.append("'{0}'".format(edu_dict[header].replace("'", "''")))
            else:
                cells.append("null")
        value_cluster = ",".join(cells)
        sql = sql_template.format(column_cluster, value_cluster)
        sqls.append(sql)
    return sqls


#file = "bolivia/bolivia_1.json"
#line = ""
#with open(file, "r", encoding="utf8") as r:
#    line = r.readline()
#    line = line.strip()
#sqls = convert_a_line(line)
#print(sqls)